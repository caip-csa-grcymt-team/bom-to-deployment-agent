"""Parse Azure Pricing Calculator Excel exports into structured data.

Extracts resource estimates from Excel workbooks exported by the Azure Pricing
Calculator, returning both raw text content and structured row data for
downstream LLM analysis.

Example:
    >>> from pathlib import Path
    >>> parser = ExcelParser()
    >>> result = parser.parse(Path("estimate.xlsx"))
    >>> print(result["structured_rows"][0])
    {'Service': 'Virtual Machines', 'Region': 'West Europe', ...}
"""

from __future__ import annotations

import logging
import re
from pathlib import Path
from typing import TYPE_CHECKING, Any

import openpyxl

if TYPE_CHECKING:
    from openpyxl.workbook import Workbook
    from openpyxl.worksheet.worksheet import Worksheet

logger = logging.getLogger(__name__)

_ESTIMATE_SHEET_PATTERNS: list[re.Pattern[str]] = [
    re.compile(r"estimate", re.IGNORECASE),
    re.compile(r"your", re.IGNORECASE),
]

_HEADER_REQUIRED_COLUMNS: set[str] = {"service", "region"}


class ExcelParserError(Exception):
    """Raised when the Excel file cannot be parsed."""


class ExcelParser:
    """Parses Azure Pricing Calculator Excel exports.

    The parser locates the estimate sheet by name heuristic, reads all rows
    as raw text, and attempts deterministic header-based extraction for
    structured row data.
    """

    def parse(self, path: Path) -> dict[str, Any]:
        """Parse an Excel workbook and return raw content with structured rows.

        Args:
            path: Path to the Excel workbook file.

        Returns:
            Dictionary with two keys:
                - ``raw_content``: All rows joined as pipe-separated text lines.
                - ``structured_rows``: List of dicts extracted from header-based
                  parsing, or an empty list if no header row is found.

        Raises:
            ExcelParserError: If the file is missing, corrupt, or contains no data.
            FileNotFoundError: If the path does not exist.
        """
        path = Path(path).resolve()
        if not path.exists():
            raise FileNotFoundError(f"Excel file not found: {path}")

        workbook = self._open_workbook(path)
        sheet = self._find_estimate_sheet(workbook)

        raw_lines = self._read_raw_lines(sheet)
        if not raw_lines:
            raise ExcelParserError(f"Workbook sheet '{sheet.title}' contains no data")

        structured_rows = self._extract_structured_rows(sheet)

        logger.info(
            "Parsed %d raw lines and %d structured rows from '%s'",
            len(raw_lines),
            len(structured_rows),
            sheet.title,
        )

        return {
            "raw_content": "\n".join(raw_lines),
            "structured_rows": structured_rows,
        }

    def _open_workbook(self, path: Path) -> Workbook:
        """Open an Excel workbook in data-only mode.

        Args:
            path: Resolved path to the workbook file.

        Returns:
            The opened openpyxl Workbook.

        Raises:
            ExcelParserError: If the file is corrupt or unreadable.
        """
        try:
            return openpyxl.load_workbook(path, data_only=True)
        except Exception as exc:
            raise ExcelParserError(f"Cannot open workbook '{path.name}': {exc}") from exc

    def _find_estimate_sheet(self, workbook: Workbook) -> Worksheet:
        """Locate the estimate sheet by name heuristic.

        Searches sheet names for patterns containing 'estimate' or 'your'
        (case-insensitive). Falls back to the active sheet when no match is found.

        Args:
            workbook: An open openpyxl Workbook.

        Returns:
            The matched or fallback Worksheet.
        """
        for pattern in _ESTIMATE_SHEET_PATTERNS:
            for name in workbook.sheetnames:
                if pattern.search(name):
                    logger.debug("Matched estimate sheet: '%s' via pattern '%s'", name, pattern.pattern)
                    return workbook[name]

        active = workbook.active
        if active is None:
            raise ExcelParserError("Workbook has no sheets")

        logger.debug("No estimate sheet matched; falling back to active sheet '%s'", active.title)
        return active

    def _read_raw_lines(self, sheet: Worksheet) -> list[str]:
        """Read all rows as pipe-separated text lines.

        Args:
            sheet: The worksheet to read.

        Returns:
            List of strings, one per row, with cell values joined by ' | '.
        """
        lines: list[str] = []
        for row in sheet.iter_rows(values_only=True):
            cell_values = [str(cell) if cell is not None else "" for cell in row]
            line = " | ".join(cell_values)
            if line.strip(" |"):
                lines.append(line)
        return lines

    def _extract_structured_rows(self, sheet: Worksheet) -> list[dict[str, Any]]:
        """Attempt header-based extraction of structured row data.

        Scans the first 10 rows for a header row that contains both 'Service'
        and 'Region' columns (case-insensitive), then reads subsequent rows as
        dictionaries keyed by header names.

        Args:
            sheet: The worksheet to extract from.

        Returns:
            List of row dicts, or empty list if no suitable header row is found.
        """
        rows = list(sheet.iter_rows(values_only=True))
        if not rows:
            return []

        header_idx, headers = self._find_header_row(rows[:10])
        if header_idx is None or headers is None:
            logger.debug("No header row with 'Service' and 'Region' found in first 10 rows")
            return []

        structured: list[dict[str, Any]] = []
        for row in rows[header_idx + 1 :]:
            values = [str(cell) if cell is not None else "" for cell in row]
            if not any(v.strip() for v in values):
                continue
            row_dict = {header: value for header, value in zip(headers, values, strict=False)}
            structured.append(row_dict)

        return structured

    def _find_header_row(self, candidate_rows: list[tuple[Any, ...]]) -> tuple[int | None, list[str] | None]:
        """Find the first row containing both 'Service' and 'Region' columns.

        Args:
            candidate_rows: Rows to scan (typically the first 10).

        Returns:
            Tuple of (row_index, header_names) or (None, None) if not found.
        """
        for idx, row in enumerate(candidate_rows):
            cell_names = [str(cell).strip().lower() if cell is not None else "" for cell in row]
            found = {col for col in _HEADER_REQUIRED_COLUMNS if col in cell_names}
            if found == _HEADER_REQUIRED_COLUMNS:
                headers = [str(cell).strip() if cell is not None else f"Column_{i}" for i, cell in enumerate(row)]
                logger.debug("Found header row at index %d: %s", idx, headers)
                return idx, headers
        return None, None
